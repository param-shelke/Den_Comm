import openpyxl
import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum, columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)


def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')


def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rownum, columnno).fill = redFill
    workbook.save(file)

# Get Login Credentials
def login_credentials(file, sheetName, row):
    lst_details = []
    for col in range(1, getColumnCount(file, sheetName) + 1):
        lst_details.append(readData(file, sheetName, row, col))
    return lst_details



# pat = r'C:\Users\Prem\Python\PycharmProjects\Pytest_Credkart - Copy\utilities\Test_datA_reg.xlsx'
# name =readData(,1,1)
# print(name)

# for row in range(1,getRowCount(pat,'data1')+1):
#     print(f"Proceesing for User{row} Credentials")
#     details = login_credentials(pat, 'data1', row)
#     name = details[0]
#     email_id = details[1]
#     password = details[2]
#     confirm_password = details[3]
#     print(name, email_id, password, confirm_password)


